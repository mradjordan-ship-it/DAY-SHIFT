"""External data sources for ticker tape: BLS wage data + RSS news feeds."""
import time
import zipfile
import io
import logging

import httpx
import feedparser

log = logging.getLogger(__name__)

# ── BLS OEWS Wage Data ──

# Restaurant industry SOC codes we care about
BLS_OCCUPATIONS = {
    "35-1011": "Chefs and Head Cooks",
    "35-2014": "Cooks, Restaurant",
    "35-2021": "Food Preparation Workers",
    "35-3021": "Fast Food & Prep Workers",
    "35-3031": "Waiters and Waitresses",
    "35-9021": "Dishwashers",
}

_bls_cache: dict = {"data": None, "timestamp": 0}
BLS_CACHE_TTL = 86400  # 24 hours — BLS data updates yearly


def fetch_bls_wages() -> list[dict]:
    """Fetch national wage data from BLS OEWS. Returns list of dicts with occupation, hourly_mean, hourly_median, employment."""
    now = time.time()
    if _bls_cache["data"] and (now - _bls_cache["timestamp"]) < BLS_CACHE_TTL:
        return _bls_cache["data"]

    try:
        r = httpx.get(
            "https://www.bls.gov/oes/special.requests/oesm23nat.zip",
            timeout=30,
            headers={"User-Agent": "DayShift/1.0 (wage data)"},
            follow_redirects=True,
        )
        if r.status_code != 200:
            log.warning("BLS zip download failed: %s", r.status_code)
            return _bls_cache["data"] or []

        import openpyxl

        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx_name = [n for n in z.namelist() if n.endswith(".xlsx")]
        if not xlsx_name:
            log.warning("No xlsx in BLS zip")
            return _bls_cache["data"] or []

        with z.open(xlsx_name[0]) as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()))
            ws = wb.active
            hdr = [cell.value for cell in ws[1]]

            results = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                occ_code = row[hdr.index("OCC_CODE")] if "OCC_CODE" in hdr else None
                if occ_code and str(occ_code) in BLS_OCCUPATIONS:
                    h_mean = row[hdr.index("H_MEAN")] if "H_MEAN" in hdr else None
                    h_median = row[hdr.index("H_MEDIAN")] if "H_MEDIAN" in hdr else None
                    tot_emp = row[hdr.index("TOT_EMP")] if "TOT_EMP" in hdr else None
                    results.append({
                        "occ_code": str(occ_code),
                        "title": BLS_OCCUPATIONS[str(occ_code)],
                        "hourly_mean": float(h_mean) if h_mean and str(h_mean) != "*" else None,
                        "hourly_median": float(h_median) if h_median and str(h_median) != "*" else None,
                        "employment": int(float(tot_emp)) if tot_emp and str(tot_emp) != "*" else None,
                    })

        _bls_cache["data"] = results
        _bls_cache["timestamp"] = now
        log.info("BLS wage data refreshed: %d occupations", len(results))
        return results

    except Exception as e:
        log.error("BLS fetch error: %s", e)
        return _bls_cache["data"] or []


# ── RSS News Feeds ──

RSS_FEEDS = [
    {
        "url": "https://www.eater.com/rss/index.xml",
        "source": "Eater",
        "icon": "📰",
    },
    {
        "url": "https://www.restaurantbusinessonline.com/feed",
        "source": "Restaurant Business",
        "icon": "📰",
    },
]

_news_cache: dict = {"data": None, "timestamp": 0}
NEWS_CACHE_TTL = 1800  # 30 minutes


def fetch_rss_news(max_per_feed: int = 5) -> list[dict]:
    """Fetch recent articles from food industry RSS feeds. Returns list of dicts."""
    now = time.time()
    if _news_cache["data"] and (now - _news_cache["timestamp"]) < NEWS_CACHE_TTL:
        return _news_cache["data"]

    all_articles = []
    for feed_info in RSS_FEEDS:
        try:
            feed = feedparser.parse(feed_info["url"])
            for entry in feed.entries[:max_per_feed]:
                title = getattr(entry, "title", "")
                link = getattr(entry, "link", "")
                summary = getattr(entry, "summary", "") or getattr(entry, "description", "")
                # Clean HTML from summary
                import re
                summary = re.sub(r"<[^>]+>", "", summary).strip()
                if len(summary) > 300:
                    summary = summary[:297] + "..."
                published = getattr(entry, "published", "") or getattr(entry, "updated", "")
                all_articles.append({
                    "title": title,
                    "link": link,
                    "summary": summary,
                    "source": feed_info["source"],
                    "icon": feed_info["icon"],
                    "published": published,
                })
        except Exception as e:
            log.error("RSS fetch error for %s: %s", feed_info["source"], e)

    _news_cache["data"] = all_articles
    _news_cache["timestamp"] = now
    log.info("RSS news refreshed: %d articles", len(all_articles))
    return all_articles
